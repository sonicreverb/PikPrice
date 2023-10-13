import time
import modules.CatalogParser as Parser
import re

from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from openpyxl import load_workbook

HOST = 'https://re-store.ru/'


# принимает на вход ссылку на страницу с категориями, возвращает массив ссылок на товары из этой категории
def get_links_from_category_page(driver, category_link):
    driver.execute_script('window.location.href = arguments[0];', category_link)
    print(f"[RE-STORE CAT.PARSER] Получение ссылок на товары из категории {category_link}")

    result_urls = []  # итоговый массив ссылок товаров на этой странице категорий

    try:
        soup = Parser.get_htmlsoup(driver)
        # некрасивая аналогия do while в python, чтобы нажать на кнопку "показать больше", пока это возможно
        if soup.find('button', class_='btn btn--black btn--size-sm btn--full-width'):
            while True:
                showMoreButton = driver.find_element(By.CSS_SELECTOR, 'button.btn--black:nth-child(1)')
                actions = ActionChains(driver)
                actions.move_to_element(showMoreButton).perform()
                showMoreButton.click()
                time.sleep(1)
                soup = Parser.get_htmlsoup(driver)
                if not soup.find('button', class_='btn btn--black btn--size-sm btn--full-width'):
                    break

        # собираем карточки продуктов, из них выдёргиваются ссылки
        product_cards = soup.find_all('div',
                                      class_="catalog__product product-card product-card--hovered product-card--touch")
        if not product_cards:
            raise Exception("Ни одна карточка товара не была найдена.")

        for product_card in product_cards:
            link = product_card.findNext('a')
            if not link:
                raise Exception(f"Не удалось получить ссылку из карточки продукта!{product_card}")

            result_urls.append(HOST + link.get('href'))

    except Exception as _ex:
        result_urls = None
        print(f'[RE-STORE CAT.PARSER] Ошибка! Во время процесса сбора ссылок было вызвано исключение!\n({_ex})')
    finally:
        print(f"[RE-STORE CAT.PARSER] Функция get_links_from_category_page вернуло значение: {result_urls}")
        return result_urls


# принимает на вход созданный драйвер, возвращает словарь product_data
def get_data(driver, url):
    driver.execute_script('window.location.href = arguments[0];', url)
    soup = Parser.get_htmlsoup(driver)

    # НАИМЕНОВАНИЕ
    name = soup.find('h1', class_="detail-heading__heading h4")
    if not name:
        raise Exception("Не удалось найти наменование продукта.")
    name = name.get_text()

    # ЦЕНА
    price = soup.find('div', class_='detail-price__current')
    if not price:
        raise Exception("Не удалось найти цену продукта.")
    price = int(price.get_text().strip()[:-1].replace(' ', ''))

    # ЦВЕТ
    color_pattern = r"«(.*?)»"
    color = re.search(color_pattern, name)
    if not color:
        # raise Exception("Не удалось найти цвет продукта.")
        color = 'NULL'
    else:
        color = color.group(1)

    # ИЗОБРАЖЕНИЯ
    try:
        images_block = soup.find('div',
                                 class_='slides-swiper__main swiper-container swiper-container-initialized '
                                        'swiper-container-horizontal')
        main_image = images_block.find('div', class_='swiper-slide swiper-slide-active').findNext('img').get('src')
        images = [elem.findNext('img').get('src') for elem in images_block.find_all('div', class_='swiper-slide')]
        images = list(set(images))

        # главное изображение всегда первое! такая заморочка из-за неадекватного расположения картинок
        mainImageInList = False
        for it in range(len(images)):
            if images[it] == main_image:
                tmp = images[it]
                images[it] = images[0]
                images[0] = tmp
                mainImageInList = True
        if not mainImageInList:
            images.insert(0, main_image)
    except Exception as _ex:
        # raise Exception(f'Не удалось найти изображения продукта\n({_ex})')
        images = []

    # ПАМЯТЬ
    ram = storage = 'NULL'
    # memories = soup.find_all('a', class_='detail-memory__link detail-memory__link--active has-tooltip')
    # memories = [memories[0]]
    # if not memories:
    #     # raise Exception("Не удалось найти память устройства")
    #     memories = ['NULL']
    # if len(memories) < 2:
    #     # raise Exception("Не удалось корректно найти память устройства.")
    #     ram = storage = 'NULL'
    # else:
    #     memories = [int(elem.get_text().strip()[:-2].replace(" ", '')) for elem in memories]
    #     ram = str(min(memories)) + 'GB'
    #     # storage = str(max(memories)) + 'GB'
    #     storage = ''

    # ОПИСАНИЕ
    try:
        showDescButton = driver.find_element(By.CSS_SELECTOR, 'li.tabs__item:nth-child(1) > button:nth-child(1)')
        actions = ActionChains(driver)
        actions.move_to_element(showDescButton).perform()
        showDescButton.click()
        time.sleep(1)

        soup = Parser.get_htmlsoup(driver)
        description_block = soup.find('div', class_='container rich content-limiter')
        description_IMG = [elem.findNext('img').get('src') for elem in description_block.find_all('picture')]
    except:
        description_IMG = []

    # ХАРАКТЕРИСТИКИ
    options = {}
    try:
        showCharButton = driver.find_element(By.CSS_SELECTOR, 'li.tabs__item:nth-child(2) > button:nth-child(1)')
        actions = ActionChains(driver)
        actions.move_to_element(showCharButton).perform()
        showCharButton.click()

        time.sleep(1)
        soup = Parser.get_htmlsoup(driver)

        table_blocks = soup.find_all('div', class_='text-small text-small--sm-min re-specs-table__row')
        for table_block in table_blocks:
            key = table_block.find('span', class_='re-specs-table__property').get_text()
            value = ''
            if table_block.find('span', class_='re-specs-table__value'):
                value = table_block.find('span', class_='re-specs-table__value').get_text()
            elif table_block.find('ul', class_='list list--specs'):
                for li in table_block.find('ul', class_='list list--specs').find_all('li'):
                    if len(li.get_text().strip()) > 1:
                        value = value + '; ' + li.get_text().strip()
            options[key.strip()] = value.strip()
    except:
        options = {}

    product_data = {'Name': name, 'Price': price, 'URL': url, 'Color': color, 'IMGs': images, 'Storage': storage,
                    'RAM': ram, 'Options': options, 'DescriptionIMGs': description_IMG}
    return product_data


def write_content_to_xlsx(product_content_dict, xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb['data']

    line_num = str(ws.max_row + 1)

    ws["A" + line_num] = product_content_dict["Name"]
    ws["B" + line_num] = product_content_dict["URL"]
    ws["C" + line_num] = product_content_dict["Price"]
    ws["D" + line_num] = product_content_dict["Storage"]
    ws["E" + line_num] = product_content_dict["RAM"]
    ws["F" + line_num] = product_content_dict["Color"]
    # ws["F" + line_num] = product_content_dict["DescHtmlRaw"]

    # G-K занимают изображения (71 - G по ascii)
    img_li = product_content_dict["IMGs"]
    for img_index in range(len(img_li)):
        if len(img_li) <= img_index:
            break
        ws[(chr(71 + img_index) + line_num)] = img_li[img_index]

    # изображения - описания
    # descriptions_img_li = product_content_dict["DescriptionIMGs"]
    # for cat_index in range(len(descriptions_img_li)):
    #     # if len(descriptions_img_li) <= cat_index:
    #     #     break
    #     ws[(chr(80 + cat_index) + line_num)] = descriptions_img_li[cat_index]

    # ХАРАКТЕРИСТИКИ
    TO_keys = []
    COLUM_NUM_TO_CHAR_CONST = 65
    for col_index in range(26):
        cell_value = str(ws['A' + chr(COLUM_NUM_TO_CHAR_CONST + col_index) + '1'].value)
        if cell_value == "None":
            break
        TO_keys.append(cell_value)

    TO_dict = product_content_dict['Options']

    for key in TO_dict:
        value = TO_dict[key]
        write_key_to_table_flag = False

        # проверка, что ключ надо записать в таблицу и значения в диапазоне до AZ1
        if key not in TO_keys and len(TO_keys) < 26:
            TO_keys.append(key)
            write_key_to_table_flag = True

        if key in TO_keys:
            index = TO_keys.index(key)
            ws["A" + chr(index + COLUM_NUM_TO_CHAR_CONST) + line_num] = value
            if write_key_to_table_flag:
                ws["A" + chr(index + COLUM_NUM_TO_CHAR_CONST) + '1'] = key
    #
    wb.save(xlsx_path)
    wb.close()


def scrapy_restore():
    CATEGORY_LINKS = ['https://re-store.ru/apple-watch/type_watchseries9/',
                      'https://re-store.ru/apple-watch/type_watchultra2/']
    driver = Parser.create_driver()
    driver.get(HOST)

    for category_link in CATEGORY_LINKS:
        prods_urls = get_links_from_category_page(driver, category_link)
        # prods_urls = ['https://re-store.ru//catalog/MQKX3/']
        if not prods_urls:
            continue

        for prod_url in prods_urls:
            try:
                data = get_data(driver, prod_url)
            except Exception as _ex:
                print(prod_url, _ex)
                data = None
            if data:
                print(data)
                write_content_to_xlsx(data, 'output.xlsx')

    Parser.kill_driver(driver)


scrapy_restore()
