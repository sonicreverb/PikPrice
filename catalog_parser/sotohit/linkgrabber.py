import requests
from bs4 import BeautifulSoup
import time
from openpyxl import load_workbook
import os
from bot.create_bot import BASE_DIR


def soto_get_links_from_page(url):
    HOST = 'https://sotohit.ru'
    html = requests.get(url)
    URL = url
    print('html 200')
    soup = BeautifulSoup(html.text, 'html.parser')
    link_list = []
    try:
        page_next = soup.find('li', {'class': 'page-next'}).find('a').get('href')
    except:
        pass
    for j in soup.find_all('div', {'class': 'product-item-name'}):
        link_list.append(HOST + j.find('a').get('href'))
    try:
        while page_next:
            html = requests.get(HOST + page_next)
            if html.status_code == 200:
                soup = BeautifulSoup(html.text, 'html.parser')
                for j in soup.find_all('div', {'class': 'product-item-name'}):
                    link_list.append(HOST + j.find('a').get('href'))
                if soup.find('li', {'class': 'page-next'}):
                    page_next = soup.find('li', {'class': 'page-next'}).find('a').get('href')
                else:
                    page_next = False
            else:
                print('error html 418')
                page_next = False
    except:
        pass
    if len(link_list) != 0 and link_list != []:
        return link_list


def sotohit_links():
    time.sleep(1)
    HOST = 'https://sotohit.ru'
    html = requests.get(HOST).text
    soup = BeautifulSoup(html, 'html.parser')

    soup_category_list = soup.find('ul', {'class': 'left_menu_menu'}).find_all('a')
    link_list = []

    for category_link in soup_category_list:
        category_link = category_link.get('href')
        if 'apple-iphone' not in category_link:
            link_list.append(soto_get_links_from_page(HOST + category_link))
    return link_list


def form_table():
    with open(os.path.join(BASE_DIR, 'bot', 'modules', 'price_parser', 'links', 'all_links (unactive).txt')) as sr:
        with open(os.path.join(BASE_DIR, 'bot', 'modules', 'price_parser', 'article_base.txt'),
                  encoding='utf-8') as inp:
            article_dict = {}
            for line in inp:
                sd = line.split()
                article_dict[sd[0]] = sd[1]
        line = 2
        wb = load_workbook(os.path.join(BASE_DIR, 'bot', 'modules', 'price_parser', 'article_base.xlsx'))
        ws = wb['data']
        for link in sr:
            try:
                html = requests.get(link.strip(), verify=False).text
                soup = BeautifulSoup(html, 'html.parser')
                title = soup.find('h1').text
                if title != 'Данная страница не найдена!' and title != 'Страница была не найдена или удалена':
                    title_spc = ''.join(title.split()).upper()
                    try:
                        ws['A' + str(line)] = title
                        ws['B' + str(line)] = title_spc
                        ws['C' + str(line)] = article_dict[title_spc]
                        ws['D' + str(line)] = link
                        print(title, title_spc, article_dict[title_spc], link)
                    except:
                        ws['A' + str(line)] = title
                        ws['B' + str(line)] = title_spc
                        ws['C' + str(line)] = 'null'
                        ws['D' + str(line)] = link
                    line += 1
            except:
                print(link)
        wb.save(os.path.join(BASE_DIR, 'bot', 'modules', 'price_parser', 'article_base.xlsx'))
        wb.close()


#form_table()
links = sotohit_links()

with open(os.path.join(BASE_DIR, 'bot', 'modules', 'price_parser', 'links', 'stomatit_links (test).txt'),
'w', encoding='utf-8') as w:
    for category_link in links:
        try:
            for link in category_link:
                w.write(link + '\n')
        except:
            pass
