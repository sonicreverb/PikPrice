import requests
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from bot.create_bot import BASE_DIR

HOST = 'https://sotohit.ru'

EXCEL_LETTERS = ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U')
CHAR1_LI = []


def get_html(url):
    r = requests.get(url)
    return r


def get_content(html, line):
    soup = BeautifulSoup(html, 'html.parser')

    title = soup.find('h1').get_text()
    price = soup.find('div', class_='price-current').get_text()[:-2]

    img_slick_track = soup.find_all('div', class_='image uk-flex-none')
    img_li = []
    for raw_img in img_slick_track:
        img_li.append(HOST + raw_img.find('img').get('src'))

    print(title, price, img_li)

    table = soup.find('table', class_='product-item-options reset-table')
    tr_even = table.find_all('th')
    tr_odd = table.find_all('td')
    tmp_list1, tmp_list2 = [], []
    for char in tr_even:
        tmp_list1.append(char.get_text().strip())
    for char in tr_odd:
        tmp_list2.append(' '.join(char.get_text().split()))

    ws['A' + str(line)] = title
    ws['B' + str(line)] = price
    for i in range(len(img_li)):
                ws[EXCEL_LETTERS[i] + str(line)] = img_li[i]
    for i in range(len(tmp_list1)):
        if tmp_list1[i] not in CHAR1_LI:
            CHAR1_LI.append(tmp_list1[i])
            ws[EXCEL_LETTERS[CHAR1_LI.index(tmp_list1[i]) + 7] + '1'] = tmp_list1[i]
        ws[EXCEL_LETTERS[CHAR1_LI.index(tmp_list1[i]) + 7] + str(line)] = tmp_list2[i]

def parse(url, line):
    html = get_html(url)
    if html.status_code == 200:
        get_content(html.text, line)
    else:
        print('error 418', url)


with open(os.path.join(BASE_DIR, 'bot', 'modules', 'price_parser', 'stomatit_links (test).txt'), encoding='utf-8') as r:
    wb = load_workbook(os.path.join(BASE_DIR, 'catalog_parser', 'soto_price.xlsx'))
    ws = wb['data']

    line = 2
    for link in r:
        try:
            parse(link.strip(), line)
            line += 1
        except:
            print(link)
    wb.save(os.path.join(BASE_DIR, 'catalog_parser', 'soto_price.xlsx'))
    wb.close()
