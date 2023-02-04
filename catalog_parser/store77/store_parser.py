import os
import re
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


def get_html(url):
    response = requests.get(url)
    if response.status_code == 200:
        return response.text
    return None


def get_content(html_text):
    soup = BeautifulSoup(html_text, "html.parser")

    # НАЗВАНИЕ, ЦЕНА, ПАМЯТЬ
    title = soup.find("h1", class_="title_card_product").get_text()
    price = "".join(soup.find("p", class_="price_title_product").get_text().strip()[:-2].split())
    try:
        memory = soup.find("a", class_="cpinfo_link active").get_text()
    except:
        memory = "None"

    # ЦВЕТ
    try:
        match = re.search("\(\D*\)", title)
        color = match[0][1:-1] if match else "Not Found"
        if color == "PRODUCT":
            color = "Product Red"
    except:
        color = "None"

    # ЦВЕТ ДЛЯ МАКБУКОВ
    if "Серый космос" in title:
        color = "Серый космос"
    elif "Серебристый" in title:
        color = "Серебристый"
    elif "Золотой" in title:
        color = "Золотой"
    elif "Старлайт" in title:
        color = "Сияющая звезда"
    elif "Темно-синий" in title:
        color = "Темно-синий"

    # ИЗОБРАЖЕНИЯ
    img_li = []
    img_div = soup.find_all("img", {"title": title})
    for img in img_div:
        img_src = img.get("src")
        if img_src and img_src not in img_li:
            img_li.append(img_src)
    img_li.pop(len(img_li) - 1)


    # КАТЕГОРИИ
    div_category = soup.find("div", class_="breadcrumbs content_lr_pad")
    rawcategory = div_category.find_all("a")
    category_li = []

    for index in range(1, len(rawcategory)):
        category_li.append(rawcategory[index].get_text())

    # ОПИСАНИЕ ХТМЛ С ТЭГАМИ
    try:
        desc_htmlraw = soup.find("div", class_="wrap_descr_b")
    except:
        desc_htmlraw = "None"

    # ХАРАКТЕРИСТИКИ
    opt_dict = dict()
    try:
        for j in soup.find_all("tr"):
            opt_dict[(j.find_all("td")[0].get_text().strip())] = j.find_all("td")[1].get_text().strip()
    except:
        pass

    # КОМЛПЕКТ
    try:
        div_package = soup.find("div", class_="package_items")
        complect_product_li = div_package.find_all("div", class_="blocks_product_fix_w card-product-package-item "
                                                                 "card-product-package-item-mobile")
        package_li = []
        for complect_prod in complect_product_li:
            package_li.append("https://store77.net" + complect_prod.find("a").get("href"))
    except:
        package_li = [None, None, None]

    ProductData = dict()
    ProductData["Title"] = title
    ProductData["Price"] = price
    ProductData["Color"] = color
    ProductData["Memory"] = memory
    ProductData["CategoryList"] = category_li
    ProductData["ImgList"] = img_li
    ProductData["StoreLink"] = line.strip()
    ProductData["OptionsDict"] = opt_dict
    ProductData["PackageLi"] = package_li
    ProductData["DescHtmlRaw"] = str(desc_htmlraw)

    return ProductData


def write_content_to_xlsx(product_content_dict, xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb['data']

    line_num = str(ws.max_row + 1)

    ws["A" + line_num] = product_content_dict["Title"]
    ws["B" + line_num] = product_content_dict["StoreLink"]
    ws["C" + line_num] = product_content_dict["Price"]
    ws["D" + line_num] = product_content_dict["Memory"]
    ws["E" + line_num] = product_content_dict["Color"]
    ws["F" + line_num] = product_content_dict["DescHtmlRaw"]

    # G-K занимают изображения (71 - G по ascii)
    img_li = product_content_dict["ImgList"]
    for img_index in range(5):
        if len(img_li) <= img_index:
            break
        ws[(chr(71 + img_index) + line_num)] = img_li[img_index]

    category_li = product_content_dict["CategoryList"]
    for cat_index in range(5):
        if len(category_li) <= cat_index:
            break
        ws[(chr(76 + cat_index) + line_num)] = category_li[cat_index]

    package_li = product_content_dict["PackageLi"]
    for pac_index in range(3):
        if len(package_li) <= pac_index:
            break
        ws[(chr(81 + pac_index) + line_num)] = package_li[pac_index]

    # ХАРАКТЕРИСТИКИ
    TO_keys = []
    COLUM_NUM_TO_CHAR_CONST = 65
    for col_index in range(26):
        cell_value = str(ws['A' + chr(COLUM_NUM_TO_CHAR_CONST + col_index) + '1'].value)
        if cell_value == "None":
            break
        TO_keys.append(cell_value)

    TO_dict = product_content_dict['OptionsDict']
    print(TO_keys)
    print(TO_dict)

    for key in TO_dict:
        value = TO_dict[key]
        write_key_to_table_flag = False

        # проверка, что ключ надо записать в таблицу и значения в диапазоне до AZ1
        if key not in TO_keys and len(TO_keys) < 26:
            TO_keys.append(key)
            write_key_to_table_flag = True

        if key in TO_keys:
            index = TO_keys.index(key)
            ws["A" + chr(index + COLUM_NUM_TO_CHAR_CONST) + line_num] = value
            if write_key_to_table_flag:
                ws["A" + chr(index + COLUM_NUM_TO_CHAR_CONST) + '1'] = key

    wb.save(xlsx_path)
    wb.close()


text_input_path = os.path.join(os.path.dirname(__file__), "links", "xiaomi", "xiaomi mobiles.txt")
xlsx_output_path = os.path.join(os.path.dirname(__file__), "xlsx catalog", "xiaomi", "xiaomi mobiles.xlsx")

line = "https://store77.net/telefony_xiaomi/telefon_xiaomi_12t_pro_8_128gb_chernyy/"
html = get_html("https://store77.net/telefony_xiaomi/telefon_xiaomi_12t_pro_8_128gb_chernyy/")
content = get_content(html)

with open(text_input_path) as r:
    for line in r:
        try:
            print(line)
            html = get_html(line.strip())
            content = get_content(html)
            write_content_to_xlsx(content, xlsx_output_path)
            print(line)
        except:
            print("error", line)


